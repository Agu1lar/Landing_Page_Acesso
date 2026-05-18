"""Import docs/ITENS.xlsx -> docs/inventario-equipamentos.csv"""
import csv
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "ITENS.xlsx"
OUT = ROOT / "inventario-equipamentos.csv"

# Specs compactas — validar com ficha da frota quando "Consulte"
AERIAL_SPECS: dict[str, tuple[str, ...]] = {
    "HB 1430": ("Tesoura", "Altura de trabalho", "~14 m", "Capacidade", "~230 kg", "Alimentação", "Elétrica"),
    "HB P830": ("Tesoura", "Altura de trabalho", "~8 m", "Capacidade", "~230 kg", "Alimentação", "Elétrica"),
    "PEP P590": ("Tesoura", "Altura de trabalho", "Consulte", "Capacidade", "Consulte", "Alimentação", "Elétrica"),
    "GS 1930s": ("Tesoura", "Altura de trabalho", "~5,8 m", "Capacidade", "~227 kg", "Alimentação", "Elétrica"),
    "AWP 30S": ("Tesoura", "Altura de trabalho", "Consulte", "Capacidade", "Consulte", "Alimentação", "Consulte"),
    "SJ III 4740": ("Tesoura", "Altura de trabalho", "~14,3 m", "Capacidade", "~227 kg", "Alimentação", "Elétrica"),
    "20 MVL": ("Tesoura", "Altura de trabalho", "Consulte", "Capacidade", "Consulte", "Alimentação", "Consulte"),
    "AM – 36": ("Tesoura", "Altura de trabalho", "Consulte", "Capacidade", "Consulte", "Alimentação", "Consulte"),
    "AM - 36": ("Tesoura", "Altura de trabalho", "Consulte", "Capacidade", "Consulte", "Alimentação", "Consulte"),
    "Z45/25j- DC": ("Lança articulada", "Alcance vertical", "~14 m", "Alcance horizontal", "~7,6 m", "Alimentação", "Diesel"),
    "Z45/25J- DC": ("Lança articulada", "Alcance vertical", "~14 m", "Alcance horizontal", "~7,6 m", "Alimentação", "Diesel"),
    "SJ III 3219": ("Tesoura", "Altura de trabalho", "~5,8 m", "Capacidade", "~227 kg", "Alimentação", "Elétrica"),
    "SJ III 3226": ("Tesoura", "Altura de trabalho", "~6,7 m", "Capacidade", "~227 kg", "Alimentação", "Elétrica"),
    "SJ III 4632": ("Tesoura", "Altura de trabalho", "~9,8 m", "Capacidade", "~227 kg", "Alimentação", "Elétrica"),
    "Z60/37 DC": ("Lança articulada", "Alcance vertical", "~18 m", "Alcance horizontal", "~11 m", "Alimentação", "Diesel"),
    "160 ATJ": ("Lança articulada", "Alcance vertical", "~16 m", "Alcance horizontal", "Consulte", "Alimentação", "Diesel"),
}

SKIP_NAMES = {
    "FERRAMENTAS",
    "ACESSORIOS",
    "ANDAIMES",
    "PLATAFORMAS AEREAS",
    "ANDAIME TUBO E BRAÇADEIRA",
    "ANDAIME TORRE",
}

FIELDS = [
    "slug",
    "nome",
    "categoria",
    "descricao_curta",
    "descricao_longa",
    "spec_1_label",
    "spec_1_valor",
    "spec_2_label",
    "spec_2_valor",
    "spec_3_label",
    "spec_3_valor",
    "spec_4_label",
    "spec_4_valor",
    "tags",
    "destaque_home",
    "disponivel",
    "entrega_obra",
    "foto_disponivel",
    "observacoes",
]


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-+", "-", text).strip("-")[:80] or "item"


def categorize_ferramenta(name: str) -> str:
    n = name.lower()
    if any(k in n for k in ("betoneira", "mangote", "vibador", "vibrador", "argamassadeira", "misturador", "bomba lameira", "bomba mangote")):
        return "concretagem"
    if "compactador" in n or "placa vibratoria" in n:
        return "compactacao"
    if any(k in n for k in ("martelete", "martelo", "rompedor")):
        return "demolicao-perfuracao"
    if any(k in n for k in ("gerador", "compressor", "transformador", "tranformador")):
        return "energia"
    if "guincho" in n or "talha" in n:
        return "outros"
    return "ferramentas-eletricas"


def resolve_aerial_specs(raw: str) -> tuple[str, ...]:
    cleaned = raw.replace("Plataforma Elevatória:", "").replace("Plataforma Elevatoria:", "").strip()
    cleaned = re.sub(r"\s*[\u2013\u2014\-]\s*Lan[cç]a Articulada.*", "", cleaned, flags=re.I).strip()
    is_artic = bool(re.search(r"lan[cç]a|Z45|Z60|ATJ", cleaned, re.I))
    for key, specs in AERIAL_SPECS.items():
        if key.replace(" ", "").lower() in cleaned.replace(" ", "").lower():
            return specs
    tipo = "Lança articulada" if is_artic else "Tesoura"
    return (tipo, "Altura de trabalho", "Consulte", "Capacidade", "Consulte", "Alimentação", "Consulte")


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    items: list[dict[str, str]] = []
    seen_slugs: dict[str, bool] = {}

    def add_item(
        nome: str,
        categoria: str,
        *,
        desc_curta: str = "",
        specs: tuple[str, ...] | None = None,
        tags: str = "",
        destaque: str = "nao",
        obs: str = "",
    ) -> None:
        nome = nome.strip()
        if not nome or nome.upper() in SKIP_NAMES or "todos os acessorios" in nome.lower():
            return
        base_slug = slugify(nome)
        slug = base_slug
        i = 2
        while slug in seen_slugs:
            slug = f"{base_slug}-{i}"
            i += 1
        seen_slugs[slug] = True
        row: dict[str, str] = {f: "" for f in FIELDS}
        row.update(
            {
                "slug": slug,
                "nome": nome,
                "categoria": categoria,
                "descricao_curta": desc_curta or f"Locação de {nome} — valores sob consulta.",
                "tags": tags,
                "destaque_home": destaque,
                "disponivel": "sim",
                "entrega_obra": "consulte",
                "foto_disponivel": "pendente",
                "observacoes": obs,
            }
        )
        if specs and len(specs) >= 7:
            row["spec_1_label"], row["spec_1_valor"] = "Tipo", specs[0]
            row["spec_2_label"], row["spec_2_valor"] = specs[1], specs[2]
            row["spec_3_label"], row["spec_3_valor"] = specs[3], specs[4]
            row["spec_4_label"], row["spec_4_valor"] = specs[5], specs[6]
        items.append(row)

    compactador_count = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        name = str(v).strip()
        if name.lower() == "compactador eletrico":
            compactador_count += 1
            if compactador_count > 1:
                name = "Compactador elétrico (unidade 2)"
        add_item(name, categorize_ferramenta(name), tags=categorize_ferramenta(name).split("-")[0])

    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, 6).value
        if v:
            add_item(str(v).strip(), "andaimes-acesso", tags="andaime")

    platform_i = 0
    for r in range(5, 25):
        v = ws.cell(r, 8).value
        if not v:
            continue
        raw = str(v).strip()
        specs = resolve_aerial_specs(raw)
        model = raw.replace("Plataforma Elevatória:", "").strip()
        nome = f"Plataforma elevatória {model}"
        platform_i += 1
        obs = "Validar specs com ficha técnica da unidade" if "Consulte" in specs else ""
        add_item(
            nome,
            "equipamentos-aereos",
            desc_curta=f"Locação de plataforma elevatória {model.split('–')[0].strip()} — sob consulta.",
            specs=specs,
            tags="plataforma,altura",
            destaque="sim" if platform_i <= 6 else "nao",
            obs=obs,
        )

    with OUT.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(items)

    cats = Counter(i["categoria"] for i in items)
    print(f"Wrote {len(items)} items to {OUT}")
    for cat, count in sorted(cats.items()):
        print(f"  {cat}: {count}")


if __name__ == "__main__":
    main()
